#IMPORTANT: I had to reference openpyxl functions on YouTube, Python and Google to know how to use them
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
# This automatically solves a column in an Excel spreadsheet

load = xl.load_workbook('transactions.xlsx')
sheet = load['Sheet1']
cell = sheet['a1']

def solve():
    # solves (multiplies by 0.9) the excel column
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row=row, column=3)
        fix = cell.value * 0.9 
        fix_cell = sheet.cell(row=row, column=4)
        fix_cell.value = fix

def create_chart():
    #creates a bar graph using values in the solved Column
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e1')

solve()
create_chart()

load.save('Solved.xlsx')
